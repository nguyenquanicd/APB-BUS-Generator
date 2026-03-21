import openpyxl
import sys
from sys import argv
import re
import os
import getpass
import shutil
import json

current_path = os.getcwd()
script_path = os.path.dirname(os.path.realpath(__file__))
APB_ADDR_WD = 16

def detect_addr_width_from_cfg(workbook):
    """Detect APB address width from CFG sheet (master/slave ADDRESS WIDTH tables)."""
    if 'CFG' not in workbook.sheetnames:
        return APB_ADDR_WD
    ws_cfg = workbook['CFG']
    widths = []

    for r in range(1, ws_cfg.max_row + 1):
        for c in range(1, ws_cfg.max_column + 1):
            cell_val = ws_cfg.cell(row=r, column=c).value
            if not (isinstance(cell_val, str) and "ADDRESS WIDTH" in cell_val.upper()):
                continue
            # Width values are listed below the "ADDRESS WIDTH" header column.
            for rr in range(r + 1, ws_cfg.max_row + 1):
                w = ws_cfg.cell(row=rr, column=c).value
                if isinstance(w, (int, float)) and int(w) > 0:
                    widths.append(int(w))

    if widths:
        return max(widths)
    return APB_ADDR_WD

def calculate_offsets(cell_value, start_address=0x0):
    units = {
        "B": 1,
        "KB": 1024,
        "MB": 1024**2,
        "GB": 1024**3
    }
    match = re.match(r"(\d+)\s*([a-zA-Z]+)", str(cell_value).strip())
    if not match:
        return None, None
        
    number = int(match.group(1))
    unit = match.group(2).upper()
    
    if unit in units:
        size_in_bytes = number * units[unit]
        end_address = start_address + size_in_bytes - 1
        
        return hex(start_address), hex(end_address)
    else:
        return None, None
class port:
    def __init__(self,name,direction, width):
        self.name = name
        self.direction = direction
        self.width = width
        self.rtl = ''
        self.port_declaration()
    def port_declaration(self):
        if self.width > 1:
            self.rtl = "  " + str(self.direction) + " [" + str(self.width-1) + ":0] " + self.name + ",\n"
        elif self.width < 1:
            print(f"ERROR: The width of port {self.name} is incorrect")
            sys.exit()
        else:
            self.rtl = "  " + str(self.direction) + " " + self.name + ",\n"
        if self.direction == 'logic':
            self.rtl = self.rtl.replace(',',';')
class master:
    def __init__(self,name):
        self.name = name
        self.slv = []
        self.slv_addr = []
    def add_slave(self,slave,addr):
        self.slv.append(slave)
        self.slv_addr.append(addr)
    def decoder(self):
        self.apb_slv()
        template = open(script_path + '/Sample/Template/decoder_template.sv', "r")
        content = template.read()
        content = content.replace("//MODULE_NAME", "m_vlsi_decoder_" + self.name.lower())
        content = content.replace("//AUTHOR",getpass.getuser())
        port_rtl_com = ""
        addr_decode_rtl = ""
        slave_select_rtl = ""
        pready_pslverr_prdata_rtl = ""
        
        # Calculate max slave name length for alignment
        max_slv_len = max([len(slv) for slv in self.slv]) if len(self.slv) > 0 else 0
        
        # Port declaration for master interface
        port_rtl_com += "  // ============================================================================\n"
        port_rtl_com += f"  // Interface of Master {self.name.upper()}\n"
        port_rtl_com += "  // ============================================================================\n"
        # port_rtl_com += "  input  logic          i_clk,\n"
        # port_rtl_com += "  input  logic          i_rstn,\n"
        port_rtl_com += f"  input  logic [{APB_ADDR_WD-1}:0]   i_paddr,\n"
        port_rtl_com += "  input  logic          i_protect_en,\n"
        port_rtl_com += "  input  logic          i_slverr_en,\n"
        port_rtl_com += "  input  logic [2:0]    i_pprot,\n"
        port_rtl_com += "  input  logic [31:0]   i_pwdata,\n"
        port_rtl_com += "  input  logic          i_pwrite,\n"
        port_rtl_com += "  input  logic          i_penable,\n"
        port_rtl_com += "  input  logic          i_psel,\n"
        port_rtl_com += "  input  logic [3:0]    i_pstrb,\n"
        port_rtl_com += "  output logic          o_pslverr,\n"
        port_rtl_com += "  output logic          o_pready,\n"
        port_rtl_com += "  output logic [31:0]   o_prdata,\n"
        
        # Port declaration for slave interfaces
        for s in range(len(self.slv)):
            slv_name = self.slv[s]
            slv_name_lower = slv_name.lower()
            slv_name_upper = slv_name.upper()
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += f"  // Interface of Slave {slv_name_upper}\n"
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += f"  output logic [{APB_ADDR_WD-1}:0]   o_paddr_{slv_name_lower},  // to {slv_name_upper}\n"
            port_rtl_com += f"  output logic          o_protect_en_{slv_name_lower},\n"
            port_rtl_com += f"  output logic          o_slverr_en_{slv_name_lower},\n"
            port_rtl_com += f"  output logic [2:0]    o_pprot_{slv_name_lower},\n"
            port_rtl_com += f"  output logic [31:0]   o_pwdata_{slv_name_lower},\n"
            port_rtl_com += f"  output logic          o_pwrite_{slv_name_lower},\n"
            port_rtl_com += f"  output logic          o_penable_{slv_name_lower},\n"
            port_rtl_com += f"  output logic          o_psel_{slv_name_lower},\n"
            port_rtl_com += f"  output logic [3:0]    o_pstrb_{slv_name_lower},\n"
            port_rtl_com += f"  input  logic          i_pslverr_{slv_name_lower},\n"
            port_rtl_com += f"  input  logic          i_pready_{slv_name_lower},\n"
            port_rtl_com += f"  input  logic [31:0]   i_prdata_{slv_name_lower},\n"
        
        # Parse per-slave address ranges once so both decode and forwarded local
        # address use the same base.
        slv_addr_ranges = []
        for s in range(len(self.slv)):
            slv_addr = self.slv_addr[s]
            # Format: "0x000000 - 0x0000FF" or single value like "0x0"
            if " - " in str(slv_addr):
                addr_parts = str(slv_addr).split(" - ")
                start_addr = addr_parts[0].strip().replace("0x", "")
                end_addr = addr_parts[1].strip().replace("0x", "")
            else:
                start_addr = str(slv_addr).strip().replace("0x", "")
                end_addr = start_addr
            slv_addr_ranges.append((start_addr, end_addr))

        # Internal select signals
        if len(self.slv) > 0:
            addr_decode_rtl = "  logic [" + str(len(self.slv)-1) + ":0] w_sel;\n\n"
            addr_decode_rtl += "  // ============================================================================\n"
            addr_decode_rtl += "  // Address Decode Logic\n"
            addr_decode_rtl += "  // ============================================================================\n"
            for s in range(len(self.slv)):
                slv_name = self.slv[s]
                slv_name_upper = slv_name.upper()
                start_addr, end_addr = slv_addr_ranges[s]
                addr_decode_rtl += f"  assign w_sel[{s}] = (i_paddr >= {APB_ADDR_WD}'h{start_addr}) & (i_paddr <= {APB_ADDR_WD}'h{end_addr});  // {slv_name_upper}\n"
        # PREADY, PSLVERR, PRDATA logic
        pready_pslverr_prdata_rtl = "  // ============================================================================\n"
        pready_pslverr_prdata_rtl += "  // PREADY, PSLVERR, PRDATA Logic\n"
        pready_pslverr_prdata_rtl += "  // ============================================================================\n"
        pready_pslverr_prdata_rtl += "  always_comb begin\n"
        pready_pslverr_prdata_rtl += "    casez (w_sel)\n"
        for s in range(len(self.slv)):
            binary_str = format(1 << s, f'0{len(self.slv)}b')
            slv_name = self.slv[s]
            slv_name_lower = slv_name.lower()
            slv_name_upper = slv_name.upper()
            pready_pslverr_prdata_rtl += f"      {len(self.slv)}'b{binary_str}: begin  // {slv_name_upper}\n"
            pready_pslverr_prdata_rtl += f"        o_pready  = i_pready_{slv_name_lower};\n"
            pready_pslverr_prdata_rtl += f"        o_pslverr = i_pslverr_{slv_name_lower};\n"
            pready_pslverr_prdata_rtl += f"        o_prdata  = i_prdata_{slv_name_lower};\n"
            pready_pslverr_prdata_rtl += "      end\n"
        pready_pslverr_prdata_rtl += "      default: begin  // Unmapped access\n"
        pready_pslverr_prdata_rtl += "        o_pready  = 1'b1;\n"
        pready_pslverr_prdata_rtl += "        o_pslverr = 1'b1;\n"
        pready_pslverr_prdata_rtl += "        o_prdata  = 32'd0;\n"
        pready_pslverr_prdata_rtl += "      end\n"
        pready_pslverr_prdata_rtl += "    endcase\n"
        pready_pslverr_prdata_rtl += "  end\n"
        
        # Slave select and signal routing
        slave_select_rtl = "  // ============================================================================\n"
        slave_select_rtl += "  // Slave Select and Signal Routing\n"
        slave_select_rtl += "  // ============================================================================\n"
        for s in range(len(self.slv)):
            slv_name = self.slv[s]
            slv_name_lower = slv_name.lower()
            slv_name_upper = slv_name.upper()
            start_addr, _ = slv_addr_ranges[s]
            slave_select_rtl += f"  // {slv_name_upper}\n"
            slave_select_rtl += f"  assign o_psel_{slv_name_lower:<{max_slv_len}}    = w_sel[{s}] & i_psel;\n"
            slave_select_rtl += f"  assign o_penable_{slv_name_lower:<{max_slv_len}} = i_penable;\n"
            slave_select_rtl += f"  assign o_pwrite_{slv_name_lower:<{max_slv_len}}  = i_pwrite;\n"
            slave_select_rtl += f"  assign o_paddr_{slv_name_lower:<{max_slv_len}}   = i_paddr - {APB_ADDR_WD}'h{start_addr};\n"
            slave_select_rtl += f"  assign o_pwdata_{slv_name_lower:<{max_slv_len}}  = i_pwdata;\n"
            slave_select_rtl += f"  assign o_pstrb_{slv_name_lower:<{max_slv_len}}   = i_pstrb;\n"
            slave_select_rtl += f"  assign o_protect_en_{slv_name_lower:<{max_slv_len}} = i_protect_en;\n"
            slave_select_rtl += f"  assign o_slverr_en_{slv_name_lower:<{max_slv_len}}  = i_slverr_en;\n"
            slave_select_rtl += f"  assign o_pprot_{slv_name_lower:<{max_slv_len}}      = i_pprot;\n\n"
        
        content = content.replace("//PORT", port_rtl_com[:-2])
        content = content.replace("//ADDRESS_DECODE", addr_decode_rtl)
        content = content.replace("//PREADY_PSLVERR_PRDATA", pready_pslverr_prdata_rtl)
        content = content.replace("//SLAVE_SELECT", slave_select_rtl)

        with open(current_path + "/RTL/Decoder/m_vlsi_decoder_" + self.name.lower() + ".sv","w") as f:
            f.write(content)
        print("Generated: RTL/Decoder/m_vlsi_decoder_" + self.name.lower() + ".sv")
    
    def apb_slv(self):
        pass

class slave:
    def __init__(self,name):
        self.name = name
        self.mst = []
        self.addr_4_mst = []
    def add_master(self,master,addr):
        self.mst.append(master)
        self.addr_4_mst.append(addr)
    def apb_mst(self):
        self.data_wd = 32
        self.addr_wd = APB_ADDR_WD
        self.port_list_mst = []
        self.port_list = []
        for i in self.mst:
            m = i.lower()
            self.port_list_mst.append (port("i_paddr" + "_" + m,"input",self.addr_wd))
            self.port_list_mst.append (port("i_protect_en" + "_" + m,"input",1))
            self.port_list_mst.append (port("i_slverr_en" + "_" + m,"input",1))
            self.port_list_mst.append (port("i_pprot" + "_" + m,"input",3))
            self.port_list_mst.append (port("i_pwdata" + "_" + m,"input",self.data_wd))
            self.port_list_mst.append (port("i_pwrite" + "_" + m,"input",1))
            self.port_list_mst.append (port("i_penable" + "_" + m,"input",1))
            self.port_list_mst.append (port("i_psel" + "_" + m,"input",1))
            self.port_list_mst.append (port("i_pstrb" + "_" + m,"input",4))
            self.port_list_mst.append (port("o_pslverr" + "_" + m,"output",1))
            self.port_list_mst.append (port("o_pready" + "_" + m,"output",1))
            self.port_list_mst.append (port("o_prdata" + "_" + m,"output",self.data_wd))

        self.port_list.append (port("i_clk" ,"input",1))
        self.port_list.append (port("i_rstn" ,"input",1))
        self.port_list.append (port("o_paddr" ,"output",self.addr_wd))
        self.port_list.append (port("o_protect_en" ,"output",1))
        self.port_list.append (port("o_slverr_en" ,"output",1))
        self.port_list.append (port("o_pprot" ,"output",3))
        self.port_list.append (port("o_pwdata" ,"output",self.data_wd))
        self.port_list.append (port("o_pwrite" ,"output",1))
        self.port_list.append (port("o_penable" ,"output",1))
        self.port_list.append (port("o_psel" ,"output",1))
        self.port_list.append (port("o_pstrb" ,"output",4))
        self.port_list.append (port("i_pslverr" ,"input",1))
        self.port_list.append (port("i_pready" ,"input",1))
        self.port_list.append (port("i_prdata" ,"input",self.data_wd))

    def arbiter(self):
        self.apb_mst()
        if len(self.mst) > 1:
            template = open(script_path + '/Sample/Template/arbiter_template.sv', "r")
            content = template.read()
            content = content.replace("//MODULE_NAME", "m_vlsi_arbiter_" + self.name.lower())
            content = content.replace("//PARAMETER", " ")
            content = content.replace("//AUTHOR",getpass.getuser())
            
            # Calculate max master name length for alignment
            max_mst_len = max([len(mst) for mst in self.mst]) if len(self.mst) > 0 else 0
            
            port_rtl_com = ""
            psel_assign_rtl = ""
            case_rtl_com = ""
            slv2mst_rtl_full = ""
            
            # Port declaration for slave interface
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += f"  // Slave Interface: {self.name.upper()}\n"
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += "  input  logic          i_clk,\n"
            port_rtl_com += "  input  logic          i_rstn,\n"
            port_rtl_com += f"  output logic [{APB_ADDR_WD-1}:0]   o_paddr,\n"
            port_rtl_com += "  output logic          o_protect_en,\n"
            port_rtl_com += "  output logic          o_slverr_en,\n"
            port_rtl_com += "  output logic [2:0]    o_pprot,\n"
            port_rtl_com += "  output logic [31:0]   o_pwdata,\n"
            port_rtl_com += "  output logic          o_pwrite,\n"
            port_rtl_com += "  output logic          o_penable,\n"
            port_rtl_com += "  output logic          o_psel,\n"
            port_rtl_com += "  output logic [3:0]    o_pstrb,\n"
            port_rtl_com += "  input  logic          i_pslverr,\n"
            port_rtl_com += "  input  logic          i_pready,\n"
            port_rtl_com += "  input  logic [31:0]   i_prdata,\n"
            
            # Port declaration for each master interface
            for mst_idx, mst_name in enumerate(self.mst):
                mst_name_lower = mst_name.lower()
                mst_name_upper = mst_name.upper()
                port_rtl_com += "  // ============================================================================\n"
                port_rtl_com += f"  // Master Interface {mst_idx+1}: {mst_name_upper}\n"
                port_rtl_com += "  // ============================================================================\n"
                port_rtl_com += f"  input  logic [{APB_ADDR_WD-1}:0]   i_paddr_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic          i_protect_en_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic          i_slverr_en_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic [2:0]    i_pprot_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic [31:0]   i_pwdata_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic          i_pwrite_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic          i_penable_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic          i_psel_{mst_name_lower},\n"
                port_rtl_com += f"  input  logic [3:0]    i_pstrb_{mst_name_lower},\n"
                port_rtl_com += f"  output logic          o_pslverr_{mst_name_lower},\n"
                port_rtl_com += f"  output logic          o_pready_{mst_name_lower},\n"
                port_rtl_com += f"  output logic [31:0]   o_prdata_{mst_name_lower},\n"
            
            # PSEL combination logic
            psel_assign_rtl = "  // ============================================================================\n"
            psel_assign_rtl += "  // Master Select Logic\n"
            psel_assign_rtl += "  // ============================================================================\n"
            psel_assign_rtl += "  assign psel_comb = {"
            for m in range(len(self.mst) - 1, -1, -1):
                if m == len(self.mst) - 1:
                    psel_assign_rtl += f"i_psel_{self.mst[m].lower()}"
                else:
                    psel_assign_rtl += f", i_psel_{self.mst[m].lower()}"
            psel_assign_rtl += "};\n\n"
            psel_assign_rtl += f"  logic [{len(self.mst)-1}:0] grant_prev;\n"
            psel_assign_rtl += "  logic          grant_handover;\n"
            psel_assign_rtl += "  logic          sel_penable;\n\n"
            psel_assign_rtl += "  always_ff @(posedge i_clk or negedge i_rstn) begin\n"
            psel_assign_rtl += "    if (!i_rstn)\n"
            psel_assign_rtl += "      grant_prev <= '0;\n"
            psel_assign_rtl += "    else\n"
            psel_assign_rtl += "      grant_prev <= grant_comb;\n"
            psel_assign_rtl += "  end\n\n"
            psel_assign_rtl += "  // Force SETUP (PENABLE=0) for one cycle when grant owner changes.\n"
            psel_assign_rtl += "  assign grant_handover = (grant_comb != grant_prev) && (|grant_comb);\n\n"

            # Case statement for master-to-slave routing
            case_rtl_com = "  // ============================================================================\n"
            case_rtl_com += "  // Master-to-Slave Data Routing\n"
            case_rtl_com += "  // ============================================================================\n"
            case_rtl_com += "  always_comb begin\n"
            case_rtl_com += "    casez (grant_comb)\n"
            for m in range(len(self.mst)):
                binary_str = format(1 << m, f'0{len(self.mst)}b')
                one_pos = binary_str.find('1')
                case_value = '?' * one_pos + binary_str[one_pos:]
                case_rtl_com += f"      {len(self.mst)}'b{case_value}: begin  // {self.mst[m].upper()}\n"
                case_rtl_com += f"        o_paddr         = i_paddr_{self.mst[m].lower()};\n"
                case_rtl_com += f"        o_protect_en    = i_protect_en_{self.mst[m].lower()};\n"
                case_rtl_com += f"        o_slverr_en     = i_slverr_en_{self.mst[m].lower()};\n"
                case_rtl_com += f"        o_pprot         = i_pprot_{self.mst[m].lower()};\n"
                case_rtl_com += f"        o_pwdata        = i_pwdata_{self.mst[m].lower()};\n"
                case_rtl_com += f"        o_pwrite        = i_pwrite_{self.mst[m].lower()};\n"
                case_rtl_com += f"        sel_penable     = i_penable_{self.mst[m].lower()};\n"
                case_rtl_com += f"        o_psel          = i_psel_{self.mst[m].lower()};\n"
                case_rtl_com += f"        o_pstrb         = i_pstrb_{self.mst[m].lower()};\n"
                case_rtl_com += "      end\n\n"
            case_rtl_com += "      default: begin\n"
            case_rtl_com += "        o_paddr         = '0;\n"
            case_rtl_com += "        o_protect_en    = '0;\n"
            case_rtl_com += "        o_slverr_en     = '0;\n"
            case_rtl_com += "        o_pprot         = '0;\n"
            case_rtl_com += "        o_pwdata        = '0;\n"
            case_rtl_com += "        o_pwrite        = '0;\n"
            case_rtl_com += "        sel_penable     = '0;\n"
            case_rtl_com += "        o_psel          = '0;\n"
            case_rtl_com += "        o_pstrb         = '0;\n"
            case_rtl_com += "      end\n"
            case_rtl_com += "    endcase\n"
            case_rtl_com += "  end\n\n"
            case_rtl_com += "  assign o_penable = o_psel ? (sel_penable & ~grant_handover) : 1'b0;\n\n"
            
            # Slave-to-master routing
            slv2mst_rtl_full = "  // ============================================================================\n"
            slv2mst_rtl_full += "  // Slave-to-Master Data Routing\n"
            slv2mst_rtl_full += "  // ============================================================================\n"
            slv2mst_rtl_full += f"  assign o_pslverr_{self.mst[0].lower():<{max_mst_len}} = grant_comb[0] ? i_pslverr : '0;\n"
            slv2mst_rtl_full += f"  assign o_pready_{self.mst[0].lower():<{max_mst_len}}  = grant_comb[0] ? i_pready  : '0;\n"
            slv2mst_rtl_full += f"  assign o_prdata_{self.mst[0].lower():<{max_mst_len}}  = grant_comb[0] ? i_prdata  : '0;\n"
            for m in range(1, len(self.mst)):
                slv2mst_rtl_full += f"  assign o_pslverr_{self.mst[m].lower():<{max_mst_len}} = grant_comb[{m}] ? i_pslverr : '0;\n"
                slv2mst_rtl_full += f"  assign o_pready_{self.mst[m].lower():<{max_mst_len}}  = grant_comb[{m}] ? i_pready  : '0;\n"
                slv2mst_rtl_full += f"  assign o_prdata_{self.mst[m].lower():<{max_mst_len}}  = grant_comb[{m}] ? i_prdata  : '0;\n"
            
            content = content.replace("//PORT", port_rtl_com[:-2])
            content = content.replace("//MASTER_NUM-1", str(len(self.mst) - 1))
            content = content.replace("//MASTER_NUM", str(len(self.mst)))
            content = content.replace("//MST2SLV_LOGIC", case_rtl_com)
            content = content.replace("//SLV2MST_LOGIC", slv2mst_rtl_full)
            content = content.replace("//PSEL", psel_assign_rtl)
            with open(current_path + "/RTL/Arbiter/m_vlsi_arbiter_" + self.name.lower() + ".sv","w") as f:
                f.write(content)
            print("Generated: RTL/Arbiter/m_vlsi_arbiter_" + self.name.lower() + ".sv")
        else:
            template = open(script_path + '/Sample/Template/arbiter_template_1vs1.sv', "r")
            content = template.read()
            content = content.replace("//MODULE_NAME", "m_vlsi_arbiter_" + self.name.lower())
            content = content.replace("//PARAMETER", " ")
            content = content.replace("//AUTHOR",getpass.getuser())
            
            max_mst_len = len(self.mst[0]) if len(self.mst) > 0 else 0
            
            assign_output_rtl = "  // ============================================================================\n"
            assign_output_rtl += "  // Direct Signal Routing (Single Master)\n"
            assign_output_rtl += "  // ============================================================================\n"
            assign_output_rtl += "  assign o_paddr         = i_paddr_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_protect_en    = i_protect_en_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_slverr_en     = i_slverr_en_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_pprot         = i_pprot_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_pwdata        = i_pwdata_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_pwrite        = i_pwrite_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_penable       = i_penable_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_psel          = i_psel_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_pstrb         = i_pstrb_%s;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_pslverr_%s = i_pslverr;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_pready_%s  = i_pready;\n" % self.mst[0].lower()
            assign_output_rtl += "  assign o_prdata_%s  = i_prdata;\n" % self.mst[0].lower()
            
            port_rtl_com = ""
            # Slave interface ports
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += f"  // Slave Interface: {self.name.upper()}\n"
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += "  input  logic          i_clk,\n"
            port_rtl_com += "  input  logic          i_rstn,\n"
            port_rtl_com += f"  output logic [{APB_ADDR_WD-1}:0]   o_paddr,\n"
            port_rtl_com += "  output logic          o_protect_en,\n"
            port_rtl_com += "  output logic          o_slverr_en,\n"
            port_rtl_com += "  output logic [2:0]    o_pprot,\n"
            port_rtl_com += "  output logic [31:0]   o_pwdata,\n"
            port_rtl_com += "  output logic          o_pwrite,\n"
            port_rtl_com += "  output logic          o_penable,\n"
            port_rtl_com += "  output logic          o_psel,\n"
            port_rtl_com += "  output logic [3:0]    o_pstrb,\n"
            port_rtl_com += "  input  logic          i_pslverr,\n"
            port_rtl_com += "  input  logic          i_pready,\n"
            port_rtl_com += "  input  logic [31:0]   i_prdata,\n"
            
            # Master interface ports
            mst_name_lower = self.mst[0].lower()
            mst_name_upper = self.mst[0].upper()
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += f"  // Master Interface: {mst_name_upper}\n"
            port_rtl_com += "  // ============================================================================\n"
            port_rtl_com += f"  input  logic [{APB_ADDR_WD-1}:0]   i_paddr_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic          i_protect_en_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic          i_slverr_en_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic [2:0]    i_pprot_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic [31:0]   i_pwdata_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic          i_pwrite_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic          i_penable_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic          i_psel_{mst_name_lower},\n"
            port_rtl_com += f"  input  logic [3:0]    i_pstrb_{mst_name_lower},\n"
            port_rtl_com += f"  output logic          o_pslverr_{mst_name_lower},\n"
            port_rtl_com += f"  output logic          o_pready_{mst_name_lower},\n"
            port_rtl_com += f"  output logic [31:0]   o_prdata_{mst_name_lower}\n"
            
            content = content.replace("//PORT", port_rtl_com)
            content = content.replace("//ASSIGN_LOGIC", assign_output_rtl)
            with open(current_path + "/RTL/Arbiter/m_vlsi_arbiter_" + self.name.lower() + ".sv","w") as f:
                f.write(content)
            print("Generated: RTL/Arbiter/m_vlsi_arbiter_" + self.name.lower() + ".sv")

            # Generate lint waiver for unused i_clk and i_rstn in 1-master arbiter
            self.generate_lint_waiver()

    def generate_lint_waiver(self):
        """Generate VC Static lint waiver for unused i_clk and i_rstn in 1-master arbiter"""
        waiver_dir = current_path + "/VC_Waiver"
        if not os.path.exists(waiver_dir):
            os.makedirs(waiver_dir)
        brk_o = "{"
        brk_c = "}"
        waiver_file = waiver_dir + f"/waiver_{self.name.lower()}.tcl"
        module_name = f"m_vlsi_arbiter_{self.name.lower()}"

        content = f"# Lint Waiver for {module_name}\n"
        content += "# Waive W240 (Input declared but not read) for i_clk and i_rstn\n"
        content += "# These ports are kept for interface consistency with multi-master arbiters\n\n"
        content += f"waive_lint -add \"W240 for {module_name} i_clk and i_rstn\" -tag W240 -filter {brk_o}Module == {module_name}{brk_c}\n"
        # content += f"waive_lint -add \"W240 for {module_name} i_rstn\" -tag W240 -filter {brk_o}Module == {module_name}{brk_c} -object i_rstn\n"

        with open(waiver_file, "w") as f:
            f.write(content)
        print("Generated: " + waiver_file)

class apb_router:
    def __init__(self, mst_list, slv_list):
        self.mst_list = mst_list
        self.slv_list = slv_list
        self.mst_names = [m.name for m in mst_list]
        self.slv_names = [s.name for s in slv_list]
    
    def generate_router(self):
        """Generate top-level APB router that instantiates all arbiters and decoders"""
        content = ""
        
        # Header
        content += "`timescale 1ns/1ps\n"
        content += "//--------------------------------------\n"
        content += "//Project: APB BUS Generator\n"
        content += "//Module: m_vlsi_apb_router\n"
        content += "//Function: APB Router Top-Level\n"
        content += "//Author: %s\n" % getpass.getuser()
        content += "//Script Author: Trthinh (Ethan), Thang Luong (superzeldalink)\n"
        content += "//Page: VLSI Technology\n"
        content += "//--------------------------------------\n\n"
        
        # Module declaration
        content += "module m_vlsi_apb_router (\n"
        
        # Global signals
        content += "  // ============================================================================\n"
        content += "  // Global Signals\n"
        content += "  // ============================================================================\n"
        content += "  input  logic          i_clk,\n"
        content += "  input  logic          i_rstn,\n\n"
        
        # Master interfaces
        max_mst_len = max([len(m) for m in self.mst_names]) if len(self.mst_names) > 0 else 0
        content += "  // ============================================================================\n"
        content += "  // Master Interfaces\n"
        content += "  // ============================================================================\n"
        for mst in self.mst_names:
            mst_lower = mst.lower()
            mst_upper = mst.upper()
            content += f"  // Master: {mst_upper}\n"
            content += f"  input  logic [{APB_ADDR_WD-1}:0]   i_paddr_{mst_lower},\n"
            content += f"  input  logic          i_protect_en_{mst_lower},\n"
            content += f"  input  logic          i_slverr_en_{mst_lower},\n"
            content += f"  input  logic [2:0]    i_pprot_{mst_lower},\n"
            content += f"  input  logic [31:0]   i_pwdata_{mst_lower},\n"
            content += f"  input  logic          i_pwrite_{mst_lower},\n"
            content += f"  input  logic          i_penable_{mst_lower},\n"
            content += f"  input  logic          i_psel_{mst_lower},\n"
            content += f"  input  logic [3:0]    i_pstrb_{mst_lower},\n"
            content += f"  output logic          o_pslverr_{mst_lower},\n"
            content += f"  output logic          o_pready_{mst_lower},\n"
            content += f"  output logic [31:0]   o_prdata_{mst_lower},\n\n"
        
        # Slave interfaces
        max_slv_len = max([len(s) for s in self.slv_names]) if len(self.slv_names) > 0 else 0
        content += "  // ============================================================================\n"
        content += "  // Slave Interfaces\n"
        content += "  // ============================================================================\n"
        for slv in self.slv_names:
            slv_lower = slv.lower()
            slv_upper = slv.upper()
            content += f"  // Slave: {slv_upper}\n"
            content += f"  output logic [{APB_ADDR_WD-1}:0]   o_paddr_{slv_lower},\n"
            content += f"  output logic          o_protect_en_{slv_lower},\n"
            content += f"  output logic          o_slverr_en_{slv_lower},\n"
            content += f"  output logic [2:0]    o_pprot_{slv_lower},\n"
            content += f"  output logic [31:0]   o_pwdata_{slv_lower},\n"
            content += f"  output logic          o_pwrite_{slv_lower},\n"
            content += f"  output logic          o_penable_{slv_lower},\n"
            content += f"  output logic          o_psel_{slv_lower},\n"
            content += f"  output logic [3:0]    o_pstrb_{slv_lower},\n"
            content += f"  input  logic          i_pslverr_{slv_lower},\n"
            content += f"  input  logic          i_pready_{slv_lower},\n"
            content += f"  input  logic [31:0]   i_prdata_{slv_lower},\n\n"
        
        # Remove trailing comma from last port
        content = content.rstrip(',\n') + '\n);\n\n'
        
        # Internal wires for interconnect
        content += "  // ============================================================================\n"
        content += "  // Internal Interconnect Wires\n"
        content += "  // ============================================================================\n"
        
        # Collect all used connections
        used_slave_conns = set()  # (slave_name) for wires used by decoder/arbiter
        used_master_conns = set()  # (master_name) for wires used by decoder/arbiter
        
        for mst in self.mst_names:
            mst_obj = next((m for m in self.mst_list if m.name == mst), None)
            if mst_obj:
                for slv in mst_obj.slv:
                    used_slave_conns.add(slv)
        
        for slv in self.slv_names:
            slv_obj = next((s for s in self.slv_list if s.name == slv), None)
            if slv_obj:
                for mst in slv_obj.mst:
                    used_master_conns.add(mst)
        
        # Wires between decoders and arbiters (per master-slave connection)
        content += "  // Decoder <-> Arbiter per master-slave connections\n"
        for mst in sorted(used_master_conns):
            mst_obj = next((m for m in self.mst_list if m.name == mst), None)
            if not mst_obj:
                continue
            mst_lower = mst.lower()
            for slv in sorted(set(mst_obj.slv)):
                slv_lower = slv.lower()
                content += f"  logic [{APB_ADDR_WD-1}:0]   w_dec_paddr_{slv_lower}_{mst_lower};\n"
                content += f"  logic          w_dec_protect_en_{slv_lower}_{mst_lower};\n"
                content += f"  logic          w_dec_slverr_en_{slv_lower}_{mst_lower};\n"
                content += f"  logic [2:0]    w_dec_pprot_{slv_lower}_{mst_lower};\n"
                content += f"  logic [31:0]   w_dec_pwdata_{slv_lower}_{mst_lower};\n"
                content += f"  logic          w_dec_pwrite_{slv_lower}_{mst_lower};\n"
                content += f"  logic          w_dec_penable_{slv_lower}_{mst_lower};\n"
                content += f"  logic          w_dec_psel_{slv_lower}_{mst_lower};\n"
                content += f"  logic [3:0]    w_dec_pstrb_{slv_lower}_{mst_lower};\n"
                content += f"  logic          w_arb_pslverr_{slv_lower}_{mst_lower};\n"
                content += f"  logic          w_arb_pready_{slv_lower}_{mst_lower};\n"
                content += f"  logic [31:0]   w_arb_prdata_{slv_lower}_{mst_lower};\n"
        content += "\n"

        # Wires for master response path from arbiter
        # Note: w_dec_mst_* wires are not needed as master signals connect directly to decoder
        for mst in sorted(used_master_conns):
            mst_lower = mst.lower()
            # Commented out unused wires - these were declared but never connected
            # Original code generated these but they were never connected:
            # - w_dec_mst_paddr_*: Master address already connects directly to decoder input
            # - w_dec_mst_protect_en_*, w_dec_mst_slverr_en_*, etc.: Not used in design
            # content += f"  logic [APB_ADDR_WD-1:0]   w_dec_mst_paddr_{mst_lower};\n"
            # content += f"  logic          w_dec_mst_protect_en_{mst_lower};\n"
            # content += f"  logic          w_dec_mst_slverr_en_{mst_lower};\n"
            # content += f"  logic [2:0]    w_dec_mst_pprot_{mst_lower};\n"
            # content += f"  logic [31:0]   w_dec_mst_pwdata_{mst_lower};\n"
            # content += f"  logic          w_dec_mst_pwrite_{mst_lower};\n"
            # content += f"  logic          w_dec_mst_penable_{mst_lower};\n"
            # content += f"  logic          w_dec_mst_psel_{mst_lower};\n"
            # content += f"  logic [3:0]    w_dec_mst_pstrb_{mst_lower};\n"
            content += f"  logic          w_arb_mst_pslverr_{mst_lower};\n"
            content += f"  logic          w_arb_mst_pready_{mst_lower};\n"
            content += f"  logic [31:0]   w_arb_mst_prdata_{mst_lower};\n"
        content += "\n"
        
        # Decoder instances
        content += "  // ============================================================================\n"
        content += "  // Decoder Instances\n"
        content += "  // ============================================================================\n"
        for mst in self.mst_names:
            mst_lower = mst.lower()
            mst_upper = mst.upper()
            mst_obj = next((m for m in self.mst_list if m.name == mst), None)
            
            content += f"  // Decoder for Master {mst_upper}\n"
            content += f"  m_vlsi_decoder_{mst_lower} u_decoder_{mst_lower} (\n"
            # content += f"    .i_clk           (i_clk),\n"
            # content += f"    .i_rstn          (i_rstn),\n"
            content += f"    .i_paddr         (i_paddr_{mst_lower}),\n"
            content += f"    .i_protect_en    (i_protect_en_{mst_lower}),\n"
            content += f"    .i_slverr_en     (i_slverr_en_{mst_lower}),\n"
            content += f"    .i_pprot         (i_pprot_{mst_lower}),\n"
            content += f"    .i_pwdata        (i_pwdata_{mst_lower}),\n"
            content += f"    .i_pwrite        (i_pwrite_{mst_lower}),\n"
            content += f"    .i_penable       (i_penable_{mst_lower}),\n"
            content += f"    .i_psel          (i_psel_{mst_lower}),\n"
            content += f"    .i_pstrb         (i_pstrb_{mst_lower}),\n"
            content += f"    .o_pslverr       (w_arb_mst_pslverr_{mst_lower}),\n"
            content += f"    .o_pready        (w_arb_mst_pready_{mst_lower}),\n"
            content += f"    .o_prdata        (w_arb_mst_prdata_{mst_lower}),\n"
            
            # Only connect slave ports that this master actually uses
            if mst_obj:
                for slv in mst_obj.slv:
                    slv_lower = slv.lower()
                    content += f"    .o_paddr_{slv_lower:<{max_slv_len}}    (w_dec_paddr_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_protect_en_{slv_lower:<{max_slv_len}} (w_dec_protect_en_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_slverr_en_{slv_lower:<{max_slv_len}}  (w_dec_slverr_en_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_pprot_{slv_lower:<{max_slv_len}}      (w_dec_pprot_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_pwdata_{slv_lower:<{max_slv_len}}     (w_dec_pwdata_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_pwrite_{slv_lower:<{max_slv_len}}     (w_dec_pwrite_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_penable_{slv_lower:<{max_slv_len}}    (w_dec_penable_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_psel_{slv_lower:<{max_slv_len}}       (w_dec_psel_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_pstrb_{slv_lower:<{max_slv_len}}      (w_dec_pstrb_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_pslverr_{slv_lower:<{max_slv_len}}    (w_arb_pslverr_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_pready_{slv_lower:<{max_slv_len}}     (w_arb_pready_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_prdata_{slv_lower:<{max_slv_len}}     (w_arb_prdata_{slv_lower}_{mst_lower}),\n"
            
            # Remove trailing comma and close
            content = content.rstrip(',\n') + '\n  );\n\n'
        
        # Arbiter instances
        content += "  // ============================================================================\n"
        content += "  // Arbiter Instances\n"
        content += "  // ============================================================================\n"
        for slv in self.slv_names:
            slv_lower = slv.lower()
            slv_upper = slv.upper()
            slv_obj = next((s for s in self.slv_list if s.name == slv), None)
            
            content += f"  // Arbiter for Slave {slv_upper}\n"
            content += f"  m_vlsi_arbiter_{slv_lower} u_arbiter_{slv_lower} (\n"
            content += f"    .i_clk           (i_clk),\n"
            content += f"    .i_rstn          (i_rstn),\n"
            content += f"    .o_paddr         (o_paddr_{slv_lower}),\n"
            content += f"    .o_protect_en    (o_protect_en_{slv_lower}),\n"
            content += f"    .o_slverr_en     (o_slverr_en_{slv_lower}),\n"
            content += f"    .o_pprot         (o_pprot_{slv_lower}),\n"
            content += f"    .o_pwdata        (o_pwdata_{slv_lower}),\n"
            content += f"    .o_pwrite        (o_pwrite_{slv_lower}),\n"
            content += f"    .o_penable       (o_penable_{slv_lower}),\n"
            content += f"    .o_psel          (o_psel_{slv_lower}),\n"
            content += f"    .o_pstrb         (o_pstrb_{slv_lower}),\n"
            content += f"    .i_pslverr       (i_pslverr_{slv_lower}),\n"
            content += f"    .i_pready        (i_pready_{slv_lower}),\n"
            content += f"    .i_prdata        (i_prdata_{slv_lower}),\n"
            
            # Only connect master ports that this slave actually uses
            if slv_obj:
                for mst in slv_obj.mst:
                    mst_lower = mst.lower()
                    content += f"    .i_paddr_{mst_lower:<{max_mst_len}}    (w_dec_paddr_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_protect_en_{mst_lower:<{max_mst_len}} (w_dec_protect_en_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_slverr_en_{mst_lower:<{max_mst_len}}  (w_dec_slverr_en_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_pprot_{mst_lower:<{max_mst_len}}      (w_dec_pprot_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_pwdata_{mst_lower:<{max_mst_len}}     (w_dec_pwdata_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_pwrite_{mst_lower:<{max_mst_len}}     (w_dec_pwrite_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_penable_{mst_lower:<{max_mst_len}}    (w_dec_penable_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_psel_{mst_lower:<{max_mst_len}}       (w_dec_psel_{slv_lower}_{mst_lower}),\n"
                    content += f"    .i_pstrb_{mst_lower:<{max_mst_len}}      (w_dec_pstrb_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_pslverr_{mst_lower:<{max_mst_len}}    (w_arb_pslverr_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_pready_{mst_lower:<{max_mst_len}}     (w_arb_pready_{slv_lower}_{mst_lower}),\n"
                    content += f"    .o_prdata_{mst_lower:<{max_mst_len}}     (w_arb_prdata_{slv_lower}_{mst_lower}),\n"
            
            # Remove trailing comma and close
            content = content.rstrip(',\n') + '\n  );\n\n'

        # Master return path outputs
        content += "  // ============================================================================\n"
        content += "  // Master Response Outputs\n"
        content += "  // ============================================================================\n"
        for mst in self.mst_names:
            mst_lower = mst.lower()
            if mst in used_master_conns:
                content += f"  assign o_pslverr_{mst_lower} = w_arb_mst_pslverr_{mst_lower};\n"
                content += f"  assign o_pready_{mst_lower}  = w_arb_mst_pready_{mst_lower};\n"
                content += f"  assign o_prdata_{mst_lower}  = w_arb_mst_prdata_{mst_lower};\n"
            else:
                content += f"  assign o_pslverr_{mst_lower} = 1'b1;\n"
                content += f"  assign o_pready_{mst_lower}  = 1'b1;\n"
                content += f"  assign o_prdata_{mst_lower}  = 32'd0;\n"
        
        content += "endmodule\n"
        
        # Write to file
        with open(current_path + "/RTL/m_vlsi_apb_router.sv", "w") as f:
            f.write(content)
        print("Generated: " + current_path + "/RTL/m_vlsi_apb_router.sv")

    def generate_tb_vcs(self):
        """Generate VCS testbench based on current master/slave topology."""
        tb_dir = current_path + "/sim/vcs"
        if not os.path.exists(tb_dir):
            os.makedirs(tb_dir)
        tb_path = tb_dir + "/tb.sv"

        def parse_start_addr(addr):
            s = str(addr).strip()
            if " - " in s:
                s = s.split(" - ")[0].strip()
            if s.lower().startswith("0x"):
                return int(s, 16)
            return int(s, 0)
        def parse_addr_range(addr):
            s = str(addr).strip()
            if " - " in s:
                lo_s, hi_s = [x.strip() for x in s.split(" - ", 1)]
                lo = int(lo_s, 16) if lo_s.lower().startswith("0x") else int(lo_s, 0)
                hi = int(hi_s, 16) if hi_s.lower().startswith("0x") else int(hi_s, 0)
                return lo, hi
            lo = int(s, 16) if s.lower().startswith("0x") else int(s, 0)
            return lo, lo
        def pick_word_addr(start_addr, end_addr, word_index):
            lo = start_addr & 0xFFFF_FFFC
            hi = end_addr & 0xFFFF_FFFC
            if hi < lo:
                hi = lo
            span_words = ((hi - lo) // 4) + 1
            idx = word_index % max(1, span_words)
            return lo + (idx * 4)

        content = ""
        content += "`timescale 1ns/1ps\n\n"
        content += "module tb;\n"
        content += f"  localparam int ADDR_W = {APB_ADDR_WD};\n"
        content += "  localparam int DATA_W = 32;\n"
        content += "  localparam int SLV_MEM_AW = 16;\n\n"
        content += "  logic clk;\n"
        content += "  logic rst_n;\n\n"

        # Master-side bus declarations
        for mst in self.mst_names:
            m = mst.lower()
            content += f"  logic [ADDR_W-1:0] {m}_paddr;\n"
            content += f"  logic [DATA_W-1:0] {m}_pwdata;\n"
            content += f"  logic [DATA_W-1:0] {m}_prdata;\n"
            content += f"  logic              {m}_pwrite;\n"
            content += f"  logic              {m}_psel;\n"
            content += f"  logic              {m}_penable;\n"
            content += f"  logic              {m}_pready;\n"
            content += f"  logic              {m}_pslverr;\n"
            content += f"  logic [3:0]        {m}_pstrb;\n"
            content += f"  logic [2:0]        {m}_pprot;\n\n"

        # Slave-side bus declarations
        for slv in self.slv_names:
            s = slv.lower()
            content += f"  logic [ADDR_W-1:0] {s}_paddr;\n"
            content += f"  logic              {s}_protect_en;\n"
            content += f"  logic              {s}_slverr_en;\n"
            content += f"  logic [2:0]        {s}_pprot;\n"
            content += f"  logic [DATA_W-1:0] {s}_pwdata;\n"
            content += f"  logic              {s}_pwrite;\n"
            content += f"  logic              {s}_penable;\n"
            content += f"  logic              {s}_psel;\n"
            content += f"  logic [3:0]        {s}_pstrb;\n"
            content += f"  logic              {s}_pslverr;\n"
            content += f"  logic              {s}_pready;\n"
            content += f"  logic [DATA_W-1:0] {s}_prdata;\n\n"

        content += "  logic [31:0] rdata;\n"
        content += "  int error_count;\n\n"
        content += "  always #5 clk = ~clk;\n\n"

        # Master BFMs
        for mst in self.mst_names:
            m = mst.lower()
            content += f"  apb_master_bfm #(.PARA_ADDR_WIDTH(ADDR_W), .PARA_DATA_WIDTH(DATA_W)) u_{m}_bfm (\n"
            content += "    .i_clk(clk), .i_rst_n(rst_n),\n"
            content += f"    .o_paddr({m}_paddr), .o_pwdata({m}_pwdata), .i_prdata({m}_prdata),\n"
            content += f"    .o_pwrite({m}_pwrite), .o_psel({m}_psel), .o_penable({m}_penable),\n"
            content += f"    .i_pready({m}_pready), .o_pstrb({m}_pstrb), .o_pprot({m}_pprot)\n"
            content += "  );\n\n"

        # DUT
        content += "  m_vlsi_apb_router u_dut (\n"
        content += "    .i_clk(clk),\n"
        content += "    .i_rstn(rst_n),\n"
        for mst in self.mst_names:
            m = mst.lower()
            content += f"    .i_paddr_{m}({m}_paddr), .i_protect_en_{m}(1'b0), .i_slverr_en_{m}(1'b0),\n"
            content += f"    .i_pprot_{m}({m}_pprot), .i_pwdata_{m}({m}_pwdata), .i_pwrite_{m}({m}_pwrite),\n"
            content += f"    .i_penable_{m}({m}_penable), .i_psel_{m}({m}_psel), .i_pstrb_{m}({m}_pstrb),\n"
            content += f"    .o_pslverr_{m}({m}_pslverr), .o_pready_{m}({m}_pready), .o_prdata_{m}({m}_prdata),\n"
        for idx, slv in enumerate(self.slv_names):
            s = slv.lower()
            comma = "," if idx != len(self.slv_names) - 1 else ""
            content += f"    .o_paddr_{s}({s}_paddr), .o_protect_en_{s}({s}_protect_en), .o_slverr_en_{s}({s}_slverr_en),\n"
            content += f"    .o_pprot_{s}({s}_pprot), .o_pwdata_{s}({s}_pwdata), .o_pwrite_{s}({s}_pwrite),\n"
            content += f"    .o_penable_{s}({s}_penable), .o_psel_{s}({s}_psel), .o_pstrb_{s}({s}_pstrb),\n"
            content += f"    .i_pslverr_{s}({s}_pslverr), .i_pready_{s}({s}_pready), .i_prdata_{s}({s}_prdata){comma}\n"
        content += "  );\n\n"

        # Slave BFMs
        for slv in self.slv_names:
            s = slv.lower()
            content += "  apb_slave_bfm #(\n"
            content += "    .PARA_ADDR_WIDTH(SLV_MEM_AW),\n"
            content += "    .PARA_DATA_WIDTH(DATA_W),\n"
            content += "    .PARA_APB_WAIT_STATES(1)\n"
            content += f"  ) u_{s}_bfm (\n"
            content += f"    .i_clk(clk), .i_rst_n(rst_n), .i_paddr({s}_paddr[SLV_MEM_AW-1:0]), .i_pwdata({s}_pwdata), .o_prdata({s}_prdata),\n"
            content += f"    .i_pwrite({s}_pwrite), .i_psel({s}_psel), .i_penable({s}_penable), .o_pready({s}_pready),\n"
            content += f"    .o_pslverr({s}_pslverr), .i_pstrb({s}_pstrb), .i_pprot({s}_pprot)\n"
            content += "  );\n\n"

        # APB protocol monitor (per slave): ACCESS must follow SETUP on same transfer.
        for slv in self.slv_names:
            s = slv.lower()
            content += f"  logic              {s}_psel_d;\n"
            content += f"  logic              {s}_penable_d;\n"
            content += f"  logic [ADDR_W-1:0] {s}_paddr_d;\n"
            content += f"  logic              {s}_pwrite_d;\n"
            content += f"  logic [3:0]        {s}_pstrb_d;\n"
            content += f"  logic [2:0]        {s}_pprot_d;\n"
            content += f"  logic [DATA_W-1:0] {s}_pwdata_d;\n"
            content += f"  always @(posedge clk or negedge rst_n) begin\n"
            content += "    if (!rst_n) begin\n"
            content += f"      {s}_psel_d    <= 1'b0;\n"
            content += f"      {s}_penable_d <= 1'b0;\n"
            content += f"      {s}_paddr_d   <= '0;\n"
            content += f"      {s}_pwrite_d  <= 1'b0;\n"
            content += f"      {s}_pstrb_d   <= '0;\n"
            content += f"      {s}_pprot_d   <= '0;\n"
            content += f"      {s}_pwdata_d  <= '0;\n"
            content += "    end else begin\n"
            content += f"      if ({s}_psel && {s}_penable && !{s}_penable_d) begin\n"
            content += f"        if (!{s}_psel_d) begin\n"
            content += f"          $display(\"[FAIL] APB setup/access violation on {s.upper()} @%0t\", $time);\n"
            content += "          error_count++;\n"
            content += "        end\n"
            content += "      end\n"
            content += f"      if ({s}_psel && {s}_penable && {s}_psel_d && {s}_penable_d) begin\n"
            content += f"        if (({s}_paddr_d !== {s}_paddr) || ({s}_pwrite_d !== {s}_pwrite) || ({s}_pstrb_d !== {s}_pstrb) || ({s}_pprot_d !== {s}_pprot)) begin\n"
            content += f"          $display(\"[FAIL] APB control changed between SETUP/ACCESS on {s.upper()} @%0t\", $time);\n"
            content += "          error_count++;\n"
            content += "        end\n"
            content += f"        if ({s}_pwrite && ({s}_pwdata_d !== {s}_pwdata)) begin\n"
            content += f"          $display(\"[FAIL] APB pwdata changed between SETUP/ACCESS on {s.upper()} @%0t\", $time);\n"
            content += "          error_count++;\n"
            content += "        end\n"
            content += "      end\n"
            content += f"      {s}_psel_d    <= {s}_psel;\n"
            content += f"      {s}_penable_d <= {s}_penable;\n"
            content += f"      {s}_paddr_d   <= {s}_paddr;\n"
            content += f"      {s}_pwrite_d  <= {s}_pwrite;\n"
            content += f"      {s}_pstrb_d   <= {s}_pstrb;\n"
            content += f"      {s}_pprot_d   <= {s}_pprot;\n"
            content += f"      {s}_pwdata_d  <= {s}_pwdata;\n"
            content += "    end\n"
            content += "  end\n\n"

        # Simple checker + tests
        content += "  task automatic check_data(input string tag, input logic [31:0] got, input logic [31:0] exp);\n"
        content += "    if (got !== exp) begin\n"
        content += "      $display(\"[FAIL] %s got=0x%08h exp=0x%08h @%0t\", tag, got, exp, $time);\n"
        content += "      error_count++;\n"
        content += "    end else begin\n"
        content += "      $display(\"[PASS] %s data=0x%08h @%0t\", tag, got, $time);\n"
        content += "    end\n"
        content += "  endtask\n\n"

        content += "  initial begin\n"
        content += "`ifdef VCS\n"
        content += "    $fsdbDumpfile(\"wave.fsdb\");\n"
        content += "    $fsdbDumpvars(0, tb, \"+all\");\n"
        content += "`endif\n"
        content += "    $dumpfile(\"wave.vcd\");\n"
        content += "    $dumpvars(0, tb);\n"
        content += "    clk = 1'b0;\n"
        content += "    rst_n = 1'b0;\n"
        content += "    error_count = 0;\n"
        content += "    repeat (5) @(posedge clk);\n"
        content += "    rst_n = 1'b1;\n"
        content += "    repeat (3) @(posedge clk);\n\n"

        # Generate one directed R/W test per valid master->slave map
        pattern = 1
        for mst in self.mst_list:
            m = mst.name.lower()
            for slv_idx, slv in enumerate(mst.slv):
                s = slv.lower()
                try:
                    start_addr = parse_start_addr(mst.slv_addr[slv_idx])
                except Exception:
                    continue
                if start_addr >= (1 << APB_ADDR_WD):
                    content += f"    // Skipped {mst.name.upper()}->{slv.upper()} test: address 0x{start_addr:08X} exceeds ADDR_W={APB_ADDR_W}\n\n"
                    continue
                # keep access word-aligned and small offset
                test_addr = start_addr & 0xFFFF_FFFC
                test_data = (0xA5000000 | ((pattern & 0xFF) << 8) | (pattern & 0xFF))
                content += f"    fork : txn_{pattern}\n"
                content += "      begin\n"
                content += f"        $display(\"Running {mst.name.upper()}->{slv.upper()} @ 0x{test_addr:08X}\");\n"
                content += f"        u_{m}_bfm.write(32'h{test_addr:08X}, 32'h{test_data:08X});\n"
                content += f"        u_{m}_bfm.read(32'h{test_addr:08X}, rdata);\n"
                content += f"        check_data(\"{mst.name.upper()}->{slv.upper()}\", rdata, 32'h{test_data:08X});\n"
                content += "      end\n"
                content += "      begin\n"
                content += "        repeat (5000) @(posedge clk);\n"
                content += f"        $display(\"[FAIL] TIMEOUT in {mst.name.upper()}->{slv.upper()} transaction\");\n"
                content += "        error_count++;\n"
                content += "      end\n"
                content += "    join_any\n"
                content += f"    disable txn_{pattern};\n\n"
                pattern += 1

        # Generate contention tests: multiple masters access same slave at same time.
        stress_iters = 8
        rr_rounds = 4
        for slv in self.slv_list:
            contenders = []  # (master_name, start_addr, end_addr)
            for mst_name in slv.mst:
                mst_obj = next((x for x in self.mst_list if x.name == mst_name), None)
                if mst_obj is None:
                    continue
                try:
                    idx = mst_obj.slv.index(slv.name)
                except ValueError:
                    continue
                try:
                    start_addr, end_addr = parse_addr_range(mst_obj.slv_addr[idx])
                except Exception:
                    continue
                contenders.append((mst_name, start_addr, end_addr))

            if len(contenders) < 2:
                continue

            # Prefer concurrent writes to different local offsets (deterministic).
            # Fall back to concurrent reads when address windows are single-address.
            c0, c1 = contenders[0], contenders[1]
            m0 = c0[0].lower()
            m1 = c1[0].lower()
            a0 = c0[1] & 0xFFFF_FFFC
            a1_base = c1[1] & 0xFFFF_FFFC
            d0 = (0xB1000000 | ((pattern & 0xFF) << 8) | 0x11)
            if (c0[1] <= c0[2]) and ((c1[1] + 4) <= c1[2]):
                a1 = (c1[1] + 4) & 0xFFFF_FFFC
                d1 = (0xB2000000 | ((pattern & 0xFF) << 8) | 0x22)
                content += f"    fork : txn_{pattern}\n"
                content += "      begin\n"
                content += f"        $display(\"Running contention {slv.name.upper()}: {c0[0].upper()}@0x{a0:08X} || {c1[0].upper()}@0x{a1:08X}\");\n"
                content += "        fork\n"
                content += f"          begin u_{m0}_bfm.write(32'h{a0:08X}, 32'h{d0:08X}); end\n"
                content += f"          begin u_{m1}_bfm.write(32'h{a1:08X}, 32'h{d1:08X}); end\n"
                content += "        join\n"
                content += f"        u_{m0}_bfm.read(32'h{a0:08X}, rdata);\n"
                content += f"        check_data(\"ARB-{slv.name.upper()}-{c0[0].upper()}\", rdata, 32'h{d0:08X});\n"
                content += f"        u_{m1}_bfm.read(32'h{a1:08X}, rdata);\n"
                content += f"        check_data(\"ARB-{slv.name.upper()}-{c1[0].upper()}\", rdata, 32'h{d1:08X});\n"
                content += "      end\n"
                content += "      begin\n"
                content += "        repeat (5000) @(posedge clk);\n"
                content += f"        $display(\"[FAIL] TIMEOUT in contention {slv.name.upper()} transaction\");\n"
                content += "        error_count++;\n"
                content += "      end\n"
                content += "    join_any\n"
                content += f"    disable txn_{pattern};\n\n"
                pattern += 1
            else:
                # Single-address window fallback: preload, then launch concurrent reads.
                d_read = (0xB4000000 | ((pattern & 0xFF) << 8) | 0x44)
                content += f"    fork : txn_{pattern}\n"
                content += "      begin\n"
                content += f"        $display(\"Running read-contention {slv.name.upper()}: {c0[0].upper()}@0x{a0:08X} || {c1[0].upper()}@0x{a1_base:08X}\");\n"
                content += f"        u_{m0}_bfm.write(32'h{a0:08X}, 32'h{d_read:08X});\n"
                content += "        fork\n"
                content += f"          begin logic [31:0] rdata_{m0}; u_{m0}_bfm.read(32'h{a0:08X}, rdata_{m0}); check_data(\"ARB-RD-{slv.name.upper()}-{c0[0].upper()}\", rdata_{m0}, 32'h{d_read:08X}); end\n"
                content += f"          begin logic [31:0] rdata_{m1}; u_{m1}_bfm.read(32'h{a1_base:08X}, rdata_{m1}); check_data(\"ARB-RD-{slv.name.upper()}-{c1[0].upper()}\", rdata_{m1}, 32'h{d_read:08X}); end\n"
                content += "        join\n"
                content += "      end\n"
                content += "      begin\n"
                content += "        repeat (5000) @(posedge clk);\n"
                content += f"        $display(\"[FAIL] TIMEOUT in read-contention {slv.name.upper()} transaction\");\n"
                content += "        error_count++;\n"
                content += "      end\n"
                content += "    join_any\n"
                content += f"    disable txn_{pattern};\n\n"
                pattern += 1

            # Intensive handover stress: repeated simultaneous writes/readbacks.
            content += f"    fork : txn_{pattern}\n"
            content += "      begin\n"
            content += f"        $display(\"Running intensive contention stress on {slv.name.upper()} ({len(contenders)} contenders)\");\n"
            for it in range(stress_iters):
                d_shared = (0xC5000000 | (((pattern + it) & 0xFF) << 8) | (it & 0xFF))
                content += "        fork\n"
                for c_idx, c in enumerate(contenders):
                    m = c[0].lower()
                    a = pick_word_addr(c[1], c[2], (it * (len(contenders) + 1)) + c_idx)
                    content += f"          begin u_{m}_bfm.write(32'h{a:08X}, 32'h{d_shared:08X}); end\n"
                content += "        join\n"
                for c_idx, c in enumerate(contenders):
                    m = c[0].lower()
                    a = pick_word_addr(c[1], c[2], (it * (len(contenders) + 1)) + c_idx)
                    content += f"        u_{m}_bfm.read(32'h{a:08X}, rdata);\n"
                    content += f"        check_data(\"STR-{slv.name.upper()}-{c[0].upper()}-{it}\", rdata, 32'h{d_shared:08X});\n"
            content += "      end\n"
            content += "      begin\n"
            content += "        repeat (50000) @(posedge clk);\n"
            content += f"        $display(\"[FAIL] TIMEOUT in intensive contention {slv.name.upper()} transaction\");\n"
            content += "        error_count++;\n"
            content += "      end\n"
            content += "    join_any\n"
            content += f"    disable txn_{pattern};\n\n"
            pattern += 1

            # All-at-once round-robin check with immediate re-request:
            # every contender starts together and issues rr_rounds write+read pairs back-to-back.
            rr_n = len(contenders)
            rr_ops = rr_rounds * 2
            content += f"    fork : txn_{pattern}\n"
            content += "      begin\n"
            content += f"        time rr_done_t_{slv.name.lower()} [0:{rr_n-1}][0:{rr_ops-1}];\n"
            content += f"        bit  rr_used_{slv.name.lower()} [0:{rr_n-1}][0:{rr_ops-1}];\n"
            content += f"        int  rr_order_{slv.name.lower()} [0:{(rr_n * rr_ops)-1}];\n"
            content += "        int rr_i;\n"
            content += "        int rr_j;\n"
            content += "        int rr_k;\n"
            content += "        int rr_r;\n"
            content += "        int rr_pos;\n"
            content += "        int rr_min_m;\n"
            content += "        int rr_min_r;\n"
            content += "        int rr_exp;\n"
            content += "        time rr_min_t;\n"
            content += f"        $display(\"Running all-at-once RR check on {slv.name.upper()} ({rr_n} contenders, {rr_rounds} rounds, immediate re-request)\");\n"
            content += f"        for (rr_i = 0; rr_i < {rr_n}; rr_i++) begin\n"
            content += f"          for (rr_k = 0; rr_k < {rr_ops}; rr_k++) begin\n"
            content += f"            rr_done_t_{slv.name.lower()}[rr_i][rr_k] = 0;\n"
            content += f"            rr_used_{slv.name.lower()}[rr_i][rr_k] = 1'b0;\n"
            content += "          end\n"
            content += "        end\n"
            content += f"        for (rr_pos = 0; rr_pos < {(rr_n * rr_ops)}; rr_pos++) begin\n"
            content += f"          rr_order_{slv.name.lower()}[rr_pos] = -1;\n"
            content += "        end\n"
            rr_data_base = (0xD5000000 | ((pattern & 0xFF) << 8))
            content += "        fork\n"
            for c_idx, c in enumerate(contenders):
                m = c[0].lower()
                a = pick_word_addr(c[1], c[2], c_idx)
                content += "          begin\n"
                content += f"            int rr_k_{m};\n"
                content += f"            int rr_wop_{m};\n"
                content += f"            int rr_rop_{m};\n"
                content += f"            logic [31:0] rr_rd_{m};\n"
                content += f"            for (rr_k_{m} = 0; rr_k_{m} < {rr_rounds}; rr_k_{m}++) begin\n"
                content += f"              rr_wop_{m} = rr_k_{m} * 2;\n"
                content += f"              rr_rop_{m} = rr_wop_{m} + 1;\n"
                content += f"              u_{m}_bfm.write(32'h{a:08X}, (32'h{rr_data_base:08X} + rr_k_{m}));\n"
                content += f"              rr_done_t_{slv.name.lower()}[{c_idx}][rr_wop_{m}] = $time;\n"
                content += f"              u_{m}_bfm.read(32'h{a:08X}, rr_rd_{m});\n"
                content += f"              check_data($sformatf(\"RRRW-{slv.name.upper()}-{c[0].upper()}-%0d\", rr_k_{m}), rr_rd_{m}, (32'h{rr_data_base:08X} + rr_k_{m}));\n"
                content += f"              rr_done_t_{slv.name.lower()}[{c_idx}][rr_rop_{m}] = $time;\n"
                content += "            end\n"
                content += "          end\n"
            content += "        join\n"
            content += f"        for (rr_pos = 0; rr_pos < {(rr_n * rr_ops)}; rr_pos++) begin\n"
            content += "          rr_min_t = 64'h7FFF_FFFF_FFFF_FFFF;\n"
            content += "          rr_min_m = -1;\n"
            content += "          rr_min_r = -1;\n"
            content += f"          for (rr_j = 0; rr_j < {rr_n}; rr_j++) begin\n"
            content += f"            for (rr_r = 0; rr_r < {rr_ops}; rr_r++) begin\n"
            content += f"              if (!rr_used_{slv.name.lower()}[rr_j][rr_r] && (rr_done_t_{slv.name.lower()}[rr_j][rr_r] < rr_min_t)) begin\n"
            content += f"                rr_min_t = rr_done_t_{slv.name.lower()}[rr_j][rr_r];\n"
            content += "                rr_min_m = rr_j;\n"
            content += "                rr_min_r = rr_r;\n"
            content += "              end\n"
            content += "            end\n"
            content += "          end\n"
            content += "          if (rr_min_m < 0) begin\n"
            content += f"            $display(\"[FAIL] RR sort failure on {slv.name.upper()} @%0t\", $time);\n"
            content += "            error_count++;\n"
            content += "          end else begin\n"
            content += f"            rr_order_{slv.name.lower()}[rr_pos] = rr_min_m;\n"
            content += f"            rr_used_{slv.name.lower()}[rr_min_m][rr_min_r] = 1'b1;\n"
            content += "          end\n"
            content += "        end\n"
            content += f"        for (rr_pos = 1; rr_pos < {(rr_n * rr_ops)}; rr_pos++) begin\n"
            content += f"          rr_exp = (rr_order_{slv.name.lower()}[rr_pos-1] + 1) % {rr_n};\n"
            content += f"          if (rr_order_{slv.name.lower()}[rr_pos] !== rr_exp) begin\n"
            content += f"            $display(\"[FAIL] RR order violation on {slv.name.upper()} at step %0d: got=%0d exp=%0d @%0t\", rr_pos, rr_order_{slv.name.lower()}[rr_pos], rr_exp, $time);\n"
            content += "            error_count++;\n"
            content += "          end\n"
            content += "        end\n"
            content += "      end\n"
            content += "      begin\n"
            content += "        repeat (50000) @(posedge clk);\n"
            content += f"        $display(\"[FAIL] TIMEOUT in all-at-once RR {slv.name.upper()} transaction\");\n"
            content += "        error_count++;\n"
            content += "      end\n"
            content += "    join_any\n"
            content += f"    disable txn_{pattern};\n\n"
            pattern += 1

            # If 3+ masters map to this slave, add a 3-way contention case too.
            if len(contenders) >= 3:
                c2 = contenders[2]
                m2 = c2[0].lower()
                a2_base = c2[1] & 0xFFFF_FFFC
                if (c2[1] + 8) <= c2[2]:
                    a1 = (c1[1] + 4) & 0xFFFF_FFFC
                    a2 = (c2[1] + 8) & 0xFFFF_FFFC
                    d1 = (0xB2000000 | ((pattern & 0xFF) << 8) | 0x22)
                    d2 = (0xB3000000 | ((pattern & 0xFF) << 8) | 0x33)
                    content += f"    fork : txn_{pattern}\n"
                    content += "      begin\n"
                    content += f"        $display(\"Running 3-way contention {slv.name.upper()}: {c0[0].upper()} || {c1[0].upper()} || {c2[0].upper()}\");\n"
                    content += "        fork\n"
                    content += f"          begin u_{m0}_bfm.write(32'h{a0:08X}, 32'h{d0:08X}); end\n"
                    content += f"          begin u_{m1}_bfm.write(32'h{a1:08X}, 32'h{d1:08X}); end\n"
                    content += f"          begin u_{m2}_bfm.write(32'h{a2:08X}, 32'h{d2:08X}); end\n"
                    content += "        join\n"
                    content += f"        u_{m0}_bfm.read(32'h{a0:08X}, rdata);\n"
                    content += f"        check_data(\"ARB3-{slv.name.upper()}-{c0[0].upper()}\", rdata, 32'h{d0:08X});\n"
                    content += f"        u_{m1}_bfm.read(32'h{a1:08X}, rdata);\n"
                    content += f"        check_data(\"ARB3-{slv.name.upper()}-{c1[0].upper()}\", rdata, 32'h{d1:08X});\n"
                    content += f"        u_{m2}_bfm.read(32'h{a2:08X}, rdata);\n"
                    content += f"        check_data(\"ARB3-{slv.name.upper()}-{c2[0].upper()}\", rdata, 32'h{d2:08X});\n"
                    content += "      end\n"
                    content += "      begin\n"
                    content += "        repeat (5000) @(posedge clk);\n"
                    content += f"        $display(\"[FAIL] TIMEOUT in 3-way contention {slv.name.upper()} transaction\");\n"
                    content += "        error_count++;\n"
                    content += "      end\n"
                    content += "    join_any\n"
                    content += f"    disable txn_{pattern};\n\n"
                    pattern += 1
                else:
                    d_read = (0xB5000000 | ((pattern & 0xFF) << 8) | 0x55)
                    content += f"    fork : txn_{pattern}\n"
                    content += "      begin\n"
                    content += f"        $display(\"Running 3-way read-contention {slv.name.upper()}: {c0[0].upper()}@0x{a0:08X} || {c1[0].upper()}@0x{a1_base:08X} || {c2[0].upper()}@0x{a2_base:08X}\");\n"
                    content += f"        u_{m0}_bfm.write(32'h{a0:08X}, 32'h{d_read:08X});\n"
                    content += "        fork\n"
                    content += f"          begin logic [31:0] rdata_{m0}; u_{m0}_bfm.read(32'h{a0:08X}, rdata_{m0}); check_data(\"ARB3-RD-{slv.name.upper()}-{c0[0].upper()}\", rdata_{m0}, 32'h{d_read:08X}); end\n"
                    content += f"          begin logic [31:0] rdata_{m1}; u_{m1}_bfm.read(32'h{a1_base:08X}, rdata_{m1}); check_data(\"ARB3-RD-{slv.name.upper()}-{c1[0].upper()}\", rdata_{m1}, 32'h{d_read:08X}); end\n"
                    content += f"          begin logic [31:0] rdata_{m2}; u_{m2}_bfm.read(32'h{a2_base:08X}, rdata_{m2}); check_data(\"ARB3-RD-{slv.name.upper()}-{c2[0].upper()}\", rdata_{m2}, 32'h{d_read:08X}); end\n"
                    content += "        join\n"
                    content += "      end\n"
                    content += "      begin\n"
                    content += "        repeat (5000) @(posedge clk);\n"
                    content += f"        $display(\"[FAIL] TIMEOUT in 3-way read-contention {slv.name.upper()} transaction\");\n"
                    content += "        error_count++;\n"
                    content += "      end\n"
                    content += "    join_any\n"
                    content += f"    disable txn_{pattern};\n\n"
                    pattern += 1

        content += "    repeat (10) @(posedge clk);\n"
        content += "    if (error_count == 0) begin\n"
        content += "      $display(\"TEST PASS\");\n"
        content += "      $finish;\n"
        content += "    end else begin\n"
        content += "      $display(\"TEST FAIL. errors=%0d\", error_count);\n"
        content += "      $fatal(1);\n"
        content += "    end\n"
        content += "  end\n\n"
        content += "  initial begin\n"
        content += "    #200000;\n"
        content += "    $fatal(1, \"TIMEOUT\");\n"
        content += "  end\n"
        content += "endmodule\n"

        with open(tb_path, "w") as f:
            f.write(content)
        print("Generated: " + tb_path)
    
    def generate_debug_json(self, input_file, sheet_name):
        """Generate debug JSON file with configuration data"""
        import datetime
        debug_data = {
            "_header": {
                "description": "APB Bus Generator Debug Information",
                "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "input_file": os.path.basename(input_file),
                "sheet_name": sheet_name,
                "author": getpass.getuser(),
                "page": "VLSI Technology"
            },
            "masters": [],
            "slaves": []
        }
        
        # Add master information
        for mst in self.mst_list:
            mst_info = {
                "name": mst.name,
                "slaves": []
            }
            for idx, slv in enumerate(mst.slv):
                mst_info["slaves"].append({
                    "name": slv,
                    "address": mst.slv_addr[idx]
                })
            debug_data["masters"].append(mst_info)
        
        # Add slave information
        for slv in self.slv_list:
            slv_info = {
                "name": slv.name,
                "masters": []
            }
            for idx, mst in enumerate(slv.mst):
                slv_info["masters"].append({
                    "name": mst,
                    "address": slv.addr_4_mst[idx]
                })
            debug_data["slaves"].append(slv_info)
        
        # Write debug JSON
        debug_dir = current_path + "/Debug"
        if not os.path.exists(debug_dir):
            os.makedirs(debug_dir)
        
        with open(debug_dir + "/debug.json", "w") as f:
            json.dump(debug_data, f, indent=2)
        print("Generated: " + debug_dir + "/debug.json")
    
    def copy_arb_rr_balance(self):
        """Copy ArbRRBalance module to RTL folder"""
        src_dir = script_path + "/Sample/ArbRRBalance"
        dst_dir = current_path + "/RTL/ArbRRBalance"
        
        if os.path.exists(src_dir):
            if os.path.exists(dst_dir):
                shutil.rmtree(dst_dir)
            shutil.copytree(src_dir, dst_dir)
            print("Copied: " + src_dir + " -> " + dst_dir)
        else:
            print("Warning: ArbRRBalance not found at " + src_dir)
    
    def generate_filelist(self):
        """Generate filelist.f with all RTL files"""
        filelist_path = current_path + "/RTL/filelist.f"
        
        # Header for filelist
        content = "#--------------------------------------\n"
        content += "# APB Bus Generator - Filelist\n"
        content += "# Generated: " + str(os.path.basename(current_path)) + "\n"
        content += "# Author: " + getpass.getuser() + "\n"
        content += "# Script Author: Trthinh (Ethan), Thang Luong (superzeldalink) \n"
        content += "# Page: VLSI Technology\n"
        content += "#--------------------------------------\n"
        
        files = []
        # Add ArbRRBalance files
        arb_rr_dir = current_path + "/RTL/ArbRRBalance"
        if os.path.exists(arb_rr_dir):
            for f in sorted(os.listdir(arb_rr_dir)):
                if f.endswith('.sv') or f.endswith('.v'):
                    files.append("$APB_BUS_GENERATOR_HOME/RTL/ArbRRBalance/" + f)
        
        # Add Arbiter files
        arbiter_dir = current_path + "/RTL/Arbiter"
        if os.path.exists(arbiter_dir):
            for f in sorted(os.listdir(arbiter_dir)):
                if f.endswith('.sv') or f.endswith('.v'):
                    files.append("$APB_BUS_GENERATOR_HOME/RTL/Arbiter/" + f)
        
        # Add Decoder files
        decoder_dir = current_path + "/RTL/Decoder"
        if os.path.exists(decoder_dir):
            for f in sorted(os.listdir(decoder_dir)):
                if f.endswith('.sv') or f.endswith('.v'):
                    files.append("$APB_BUS_GENERATOR_HOME/RTL/Decoder/" + f)
        
        # Add top-level router
        files.append("$APB_BUS_GENERATOR_HOME/RTL/m_vlsi_apb_router.sv")
        
        # Write header and files
        with open(filelist_path, "w") as f:
            f.write(content)
            for file_path in files:
                f.write(file_path + "\n")
        print("Generated: " + filelist_path)
    
    def generate_source_me(self):
        """Generate source_me.sh file"""
        source_me_path = current_path + "/source_me.sh"
        
        content = "#!/bin/bash\n"
        content += "#--------------------------------------\n"
        content += "# APB Bus Generator - Source Me File\n"
        content += "#--------------------------------------\n\n"
        content += "# Set the APB Bus Generator home directory\n"
        content += "export APB_BUS_GENERATOR_HOME=\"" + current_path + "\"\n\n"
        content += "# Add RTL directory to search paths\n"
        content += "export APB_BUS_GENERATOR_RTL=\"$APB_BUS_GENERATOR_HOME/RTL\"\n\n"
        content += "# Source the filelist for simulation/synthesis tools\n"
        content += "export APB_BUS_GENERATOR_FILELIST=\"$APB_BUS_GENERATOR_RTL/filelist.f\"\n\n"
        content += "# Print status\n"
        content += "echo \"APB Bus Generator Environment Loaded\"\n"
        content += "echo \"  APB_BUS_GENERATOR_HOME = $APB_BUS_GENERATOR_HOME\"\n"
        content += "echo \"  APB_BUS_GENERATOR_RTL = $APB_BUS_GENERATOR_RTL\"\n"
        content += "echo \"  Filelist = $APB_BUS_GENERATOR_FILELIST\"\n"
        
        with open(source_me_path, "w") as f:
            f.write(content)
        
        # Make executable
        os.chmod(source_me_path, 0o755)

        print("Generated: " + source_me_path)
    
    def copy_readme(self):
        """Copy README.md from script directory to current path"""
        src_path = script_path + "/README.md"
        dst_path = current_path + "/README.md"
        
        if os.path.exists(src_path):
            shutil.copy2(src_path, dst_path)
            print("Generated: " + dst_path)
        else:
            print("Warning: README.md not found at " + src_path)

def main(args=None):
    global APB_ADDR_WD

    if args is None:
        args = argv[1:]

    if len(args) < 2:
        print("Usage: python3 apb_bus_generator.py <input.xlsx> <sheet> [all|rtl|tb]")
        return 1

    input_file = args[0]
    sheet_name = args[1]
    mode = args[2].lower() if len(args) > 2 else "all"
    if mode not in ["all", "rtl", "tb"]:
        print(f"ERROR: Unsupported mode '{mode}'. Use one of: all, rtl, tb")
        return 1

    wb = openpyxl.load_workbook(input_file, data_only=True)
    APB_ADDR_WD = detect_addr_width_from_cfg(wb)
    ws = wb[sheet_name]

    # Create output directories first
    if not os.path.exists(current_path + "/RTL"):
        os.makedirs(current_path + "/RTL")
    if not os.path.exists(current_path + "/RTL/Arbiter"):
        os.makedirs(current_path + "/RTL/Arbiter")
    if not os.path.exists(current_path + "/RTL/Decoder"):
        os.makedirs(current_path + "/RTL/Decoder")
    if not os.path.exists(current_path + "/Debug"):
        os.makedirs(current_path + "/Debug")

    # Find the row with master names (contains "Slave / Master" or similar)
    start_row = 1
    for row in range(1, ws.max_row + 1):
        first_cell = ws.cell(row=row, column=1).value
        if first_cell and "Slave" in str(first_cell) and "Master" in str(first_cell):
            start_row = row
            break

    # Find master names from the header row
    mst_row = start_row
    slv_col = 1
    mst_list = []
    slv_list = []
    slv_name_list = []

    # Get master names from header row (columns 2 onwards)
    master_names = []
    for col in range(2, ws.max_column + 1):
        mst_name = ws.cell(row=mst_row, column=col).value
        if mst_name and mst_name != "None":
            master_names.append(mst_name)

    # Create master objects
    for mst_name in master_names:
        mst_obj = master(mst_name)
        mst_list.append(mst_obj)

    # Process each slave row
    for row in range(start_row + 1, ws.max_row + 1):
        slv_name = ws.cell(row=row, column=slv_col).value
        if slv_name and slv_name != "None":
            slv_object = slave(slv_name)

            # For each master, check if slave is mapped
            for col_idx, mst_name in enumerate(master_names, start=2):
                slv_addr = ws.cell(row=row, column=col_idx).value
                if slv_addr and str(slv_addr).strip().upper() not in ["NONE", "X"]:
                    # Add slave to master
                    for mst_obj in mst_list:
                        if mst_obj.name == mst_name:
                            mst_obj.add_slave(slv_name, slv_addr)
                            break
                    # Add master to slave
                    slv_object.add_master(mst_name, slv_addr)

            if slv_name not in slv_name_list:
                slv_list.append(slv_object)
                slv_name_list.append(slv_name)

    if mode in ["all", "rtl"]:
        for s in slv_list:
            s.arbiter()

        # Generate decoders using MM-INFO sheet if it exists (has complete address ranges)
        if 'MM-INFO' in wb.sheetnames:
            ws_info = wb['MM-INFO']
            # Find the row with master names
            start_row_info = 1
            for row in range(1, ws_info.max_row + 1):
                first_cell = ws_info.cell(row=row, column=1).value
                if first_cell and "Slave" in str(first_cell) and "Master" in str(first_cell):
                    start_row_info = row
                    break

            # Get master names from header row
            master_names_info = []
            for col in range(2, ws_info.max_column + 1):
                mst_name = ws_info.cell(row=start_row_info, column=col).value
                if mst_name and mst_name != "None":
                    master_names_info.append(mst_name)

            # Create master objects for decoder
            mst_decoder_list = []
            for mst_name in master_names_info:
                mst_obj = master(mst_name)
                mst_decoder_list.append(mst_obj)

            # Process each slave row for address ranges
            for row in range(start_row_info + 1, ws_info.max_row + 1):
                slv_name = ws_info.cell(row=row, column=1).value
                if slv_name and slv_name != "None":
                    for col_idx, mst_name in enumerate(master_names_info, start=2):
                        slv_addr_range = ws_info.cell(row=row, column=col_idx).value
                        if slv_addr_range and slv_addr_range != "None" and slv_addr_range != "X":
                            for mst_obj in mst_decoder_list:
                                if mst_obj.name == mst_name:
                                    mst_obj.add_slave(slv_name, slv_addr_range)
                                    break

            for m in mst_decoder_list:
                m.decoder()

    # Generate top-level router object
    router = apb_router(mst_list, slv_list)

    if mode in ["all", "rtl"]:
        router.generate_router()
        router.generate_debug_json(input_file, sheet_name)
        router.copy_arb_rr_balance()
        router.generate_filelist()
        router.generate_source_me()

    if mode in ["all", "tb"]:
        router.generate_tb_vcs()

    return 0


if __name__ == "__main__":
    sys.exit(main())








        
